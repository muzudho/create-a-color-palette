import openpyxl as xl
#import random

from openpyxl.styles import PatternFill

from src.create_color_pallete import Color, ToneSystem


class OutputGradation():


    def play(number_of_color_samples, start_hue, saturation, brightness, exshell):
        # ãƒ¯ãƒ¼ã‚¯ãƒ–ãƒƒã‚¯ã‚’æ–°è¦ç”Ÿæˆ
        wb = xl.Workbook()

        # ãƒ¯ãƒ¼ã‚¯ã‚·ãƒ¼ãƒˆ
        ws = wb['Sheet']

        low, high = OutputGradation.create_tone(
                saturation=saturation,
                brightness=brightness)
        
        # è‰²ç›¸ [0.0, 1.0]
        #cur_hue = random.uniform(0, 1)
        cur_hue = start_hue
        step_hue = 1 / number_of_color_samples
#     print(f"""\
# {step_hue=}""")


        cell = ws[f'A1']
        cell.value = "No"

        cell = ws[f'B1']
        cell.value = "è‰²"

        cell = ws[f'C1']
        cell.value = "ã‚¦ã‚§ãƒ–ãƒ»ã‚»ãƒ¼ãƒ•ãƒ»ã‚«ãƒ©ãƒ¼"

        # ãƒ‡ãƒãƒƒã‚°ç”¨æƒ…å ±
        # cell = ws[f'A1']
        # cell.value = "è‰²ç›¸"

        # cell = ws[f'B1']
        # cell.value = "è‰²ç›¸ç¨®é¡"

        # cell = ws[f'C1']
        # cell.value = "è‰²ç›¸å†…æ®µéš"


        for index, row_th in enumerate(range(2, 2 + number_of_color_samples)):

            tone_system = ToneSystem(
                    low=low,
                    high=high,
                    hue=cur_hue)

            color_obj = Color(tone_system.get_red(), tone_system.get_green(), tone_system.get_blue())
    
            web_safe_color = color_obj.to_web_safe_color()
            xl_color = web_safe_color[1:]
            try:
                pattern_fill = PatternFill(
                        patternType='solid',
                        fgColor=xl_color)
            except:
                print(f'{xl_color=}')
                raise

            # é€£ç•ª
            cell = ws[f'A{row_th}']
            cell.value = index

            # è‰²
            cell = ws[f'B{row_th}']
            cell.fill = pattern_fill

            # ã‚¦ã‚§ãƒ–ãƒ»ã‚»ãƒ¼ãƒ•ãƒ»ã‚«ãƒ©ãƒ¼
            cell = ws[f'C{row_th}']
            cell.value = web_safe_color

            # ãƒ‡ãƒãƒƒã‚°æƒ…å ±
            # cell = ws[f'A{row_th}']
            # cell.value = cur_hue

            # cell = ws[f'B{row_th}']
            # cell.value = tone_system.get_phase_name()

            # cell = ws[f'C{row_th}']
            # cell.value = tone_system.get_value_of_hue_in_phase()

            cur_hue += step_hue
            if 1 < cur_hue:
                cur_hue -= 1


        # ãƒ¯ãƒ¼ã‚¯ãƒ–ãƒƒã‚¯ä¿å­˜
        exshell.save_workbook(wb=wb)


        is_successful = False

        # ã‚¨ã‚¯ã‚»ãƒ«é–‹ã
        exshell.open_virtual_display()


        message = f"""\
ğŸ™‹ã€€Please input
-----------------
ã‚°ãƒ©ãƒ‡ãƒ¼ã‚·ãƒ§ãƒ³ã‚’ä½œæˆã—ã¾ã—ãŸã€‚

ã‚¢ãƒ—ãƒªã‚±ãƒ¼ã‚·ãƒ§ãƒ³ã‚’çµ‚äº†ã™ã‚‹ãªã‚‰ y ã‚’ã€
ã‚„ã‚Šç›´ã™å ´åˆã¯ã€€ãã‚Œä»¥å¤–ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚

    Example of input
    ----------------
    y

Input
-----
"""
        line = input(message)
        print() # ç©ºè¡Œ


        # ã‚¨ã‚¯ã‚»ãƒ«é–‰ã˜ã‚‹
        exshell.close_virtual_display()


        return line == 'y'


    @staticmethod
    def create_tone(saturation, brightness):
        """è‰²èª¿ã‚’ï¼‘ã¤ã«æ±ºã‚ã¾ã™ã€‚

        Parameters
        ----------
        saturation : int
            å½©åº¦ã€‚[0, 255] ã®æ•´æ•°
            NOTE ãƒ¢ãƒã‚¯ãƒ­ã«è¿‘ã¥ãã¨ã€æ¨™æœ¬æ•°ãŒå¤šããªã‚‹ã¨ã€è‰²ã®é•ã„ã‚’å‡ºã—ã«ãã„ã§ã™ã€‚
        brightness : int
            æ˜åº¦
        """

        # NOTE ã‚¦ã‚§ãƒ–ãƒ»ã‚»ãƒ¼ãƒ•ãƒ»ã‚«ãƒ©ãƒ¼ã¯ã€æš—ã„è‰²ã®å¹…ãŒå¤šã‚ã«å–ã‚‰ã‚Œã¦ã„ã¾ã™ã€‚ 0ï½255 ã®ã†ã¡ã€ 180 ãã‚‰ã„ã¾ã§æš—ã„è‰²ã§ã™ã€‚
        # NOTE è‰²ã®æ¨™æœ¬æ•°ãŒå¤šããªã‚‹ã¨ã€ low, high ã¯æ¥µç«¯ã«ã§ãã¾ã›ã‚“ã€‚å¤‰åŒ–ã®å¹…ãŒç‹­ã¾ã£ã¦ã—ã¾ã„ã¾ã™ã€‚

        # ä¸Šé™
        high = brightness
        # ä¸‹é™
        low = saturation

        if 255 < high:
            raise ValueError(f'{high=} Others: {brightness=} {saturation=}')

        if low < 0:
            raise ValueError(f'{low=} Others: {brightness=} {saturation=}')


        return low, high
