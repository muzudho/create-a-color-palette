import openpyxl as xl
import subprocess
import time

from openpyxl.styles import PatternFill


class InputNumberOfColorsYouWantToCreate():


    def play(abs_path_to_contents, excel_application_path):
        message = """\
🙋　Please input
-----------------
作りたい色の数を 1 以上、常識的な数以下の整数で入力してください。

    Guide
    -----
    *   `3` - ３色
    * `100` - １００色

    Example of input
    ----------------
    7

Input
-----
"""
        line = input(message)
        number_of_color_samples = int(line)
        print() # 空行


        # ワークブックを新規生成
        wb = xl.Workbook()

        # ワークシート
        ws = wb['Sheet']

        cell = ws[f'A1']
        cell.value = "No"

        cell = ws[f'B1']
        cell.value = "色"


        for index, row_th in enumerate(range(2, 2 + number_of_color_samples)):

            web_safe_color = '#FFFFFF'
            xl_color = web_safe_color[1:]
            try:
                pattern_fill = PatternFill(
                        patternType='solid',
                        fgColor=xl_color)
            except:
                print(f'{xl_color=}')
                raise

            # 連番
            cell = ws[f'A{row_th}']
            cell.value = index

            # 色
            cell = ws[f'B{row_th}']
            cell.fill = pattern_fill

            # コメント
            cell = ws[f'C{row_th}']
            cell.value = '未定'


        print(f"""\
Save 📄［ {abs_path_to_contents} ］ contents file...
""")
        wb.save(abs_path_to_contents)


        # エクセルを開く
        print(f"""\
🔧　Open Excel...
""")
        opened_excel_process = subprocess.Popen([excel_application_path, abs_path_to_contents])   # Excel が開くことを期待
        time.sleep(1)


        message = f"""\
🙋　Please input
-----------------
開いたワークシートは、サンプルです。
次に進むために、こちらに何も文字を入力せずエンターキーを入力してください。

    Example of input
    ----------------
    

Input
-----
"""
        line = input(message)
        print() # 空行


        # エクセルを閉じる
        print(f"""\
🔧　Close Excel...
""")
        opened_excel_process.terminate()
        time.sleep(1)


        return number_of_color_samples
