import openpyxl as xl
import math
import random
import traceback

from openpyxl.styles import PatternFill
from src.create_color_pallete import Color, ToneSystem


MAX_scalar = 255


def main():
    message = """\
ä½œã‚ŠãŸã„è‰²ã®æ•°ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚

Example
-------
3

Input
-----
"""
    line = input(message)
    number_of_color_samples = int(line)


    message = f"""\
å½©åº¦ã‚’ 0 ä»¥ä¸Š {MAX_scalar} ä»¥ä¸‹ã®æ•´æ•°ã§å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚
0 ã«è¿‘ã„ã»ã©ã‚°ãƒ¬ãƒ¼ã€{MAX_scalar} ã«è¿‘ã„ã»ã©ãƒ“ãƒ“ãƒƒãƒ‰ã«è¿‘ã¥ãã¾ã™ã€‚

Example
-------
180

Input
-----
"""
    line = input(message)
    saturation = int(line)


    message = f"""\
æ˜åº¦ã‚’ {saturation} ä»¥ä¸Š {MAX_scalar} ä»¥ä¸‹ã®æ•´æ•°ã§å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚
0 ã«è¿‘ã„ã»ã©é»’ã€{MAX_scalar} ã«è¿‘ã„ã»ã©ç™½ã«è¿‘ã¥ãã¾ã™ã€‚

Example
-------
{saturation}

Input
-----
"""
    line = input(message)
    brightness = int(line)


    # ãƒ¯ãƒ¼ã‚¯ãƒ–ãƒƒã‚¯ã‚’æ–°è¦ç”Ÿæˆ
    wb = xl.Workbook()

    # ãƒ¯ãƒ¼ã‚¯ã‚·ãƒ¼ãƒˆ
    ws = wb['Sheet']

    low, high = create_tone(
            saturation=saturation,
            brightness=brightness)
    
    # è‰²ç›¸ [0.0, 1.0]
    cur_hue = random.uniform(0, 1)
    step_hue = 1 / number_of_color_samples
#     print(f"""\
# {step_hue=}""")


    cell = ws[f'A1']
    cell.value = "è‰²ç›¸"

    cell = ws[f'B1']
    cell.value = "è‰²ç›¸ç¨®é¡"

    cell = ws[f'C1']
    cell.value = "è‰²ç›¸å†…æ®µéš"

    cell = ws[f'D1']
    cell.value = "è‰²"

    cell = ws[f'E1']
    cell.value = "ã‚³ãƒ¼ãƒ‰"


    for row_th in range(2, 2 + number_of_color_samples):

        tone_system = ToneSystem(
                low=low,
                high=high,
                hue=cur_hue)
#         print(f"""\
# {low=}
# {high=}
# {cur_hue=}""")

        color_obj = Color(tone_system.get_red(), tone_system.get_green(), tone_system.get_blue())
#         print(f"""\
# {color_obj.to_web_safe_color()=}""")
    
        web_safe_color = color_obj.to_web_safe_color()
        xl_color = web_safe_color[1:]
        try:
            pattern_fill = PatternFill(
                    patternType='solid',
                    fgColor=xl_color)
        except:
            print(f'{xl_color=}')
            raise

        cell = ws[f'A{row_th}']
        cell.value = cur_hue

        cell = ws[f'B{row_th}']
        cell.value = tone_system.get_phase_name()

        cell = ws[f'C{row_th}']
        cell.value = tone_system.get_value_of_hue_in_phase()

        cell = ws[f'D{row_th}']
        cell.fill = pattern_fill

        cell = ws[f'E{row_th}']
        cell.value = web_safe_color

        cur_hue += step_hue
        if 1 < cur_hue:
            cur_hue -= 1


    file_path_to_write = './temp/hello.xlsx'
    wb.save(file_path_to_write)
    print(f"Please look ğŸ“„ï¼»{file_path_to_write}ï¼½ file.")


def create_tone(saturation, brightness):
    """è‰²èª¿ã‚’ï¼‘ã¤ã«æ±ºã‚ã¾ã™ã€‚

    Parameters
    ----------
    saturation : int
        å½©åº¦ã€‚[0, 255] ã®æ•´æ•°
        NOTE ãƒ¢ãƒã‚¯ãƒ­ã«è¿‘ã¥ãã¨ã€æ¨™æœ¬æ•°ãŒå¤šããªã‚‹ã¨ã€è‰²ã®é•ã„ã‚’å‡ºã—ã«ãã„ã§ã™ã€‚
    brightness : int
        æ˜åº¦
    """

    # NOTE ã‚¦ã‚§ãƒ–ãƒ»ã‚»ãƒ¼ãƒ•ãƒ»ã‚«ãƒ©ãƒ¼ã¯ã€æš—ã„è‰²ã®å¹…ãŒå¤šã‚ã«å–ã‚‰ã‚Œã¦ã„ã¾ã™ã€‚ 0ï½255 ã®ã†ã¡ã€ 180 ãã‚‰ã„ã¾ã§æš—ã„è‰²ã§ã™ã€‚
    # NOTE è‰²ã®æ¨™æœ¬æ•°ãŒå¤šããªã‚‹ã¨ã€ low, high ã¯æ¥µç«¯ã«ã§ãã¾ã›ã‚“ã€‚å¤‰åŒ–ã®å¹…ãŒç‹­ã¾ã£ã¦ã—ã¾ã„ã¾ã™ã€‚

    # ä¸Šé™
    high = brightness
    # ä¸‹é™
    low = brightness - saturation

    if 255 < high:
        raise ValueError(f'{high=} Others: {brightness=} {saturation=}')

    if low < 0:
        raise ValueError(f'{low=} Others: {brightness=} {saturation=}')


    return low, high


##########################
# MARK: ã‚³ãƒãƒ³ãƒ‰ã‹ã‚‰å®Ÿè¡Œæ™‚
##########################

if __name__ == '__main__':
    try:
        main()

    except Exception as err:
        print(f"""\
ãŠãŠã€æ®‹å¿µï¼ã€€ä¾‹å¤–ãŒæŠ•ã’ã‚‰ã‚Œã¦ã—ã¾ã£ãŸï¼
{type(err)=}  {err=}

ä»¥ä¸‹ã¯ã‚¹ã‚¿ãƒƒã‚¯ãƒˆãƒ¬ãƒ¼ã‚¹è¡¨ç¤ºã˜ã‚ƒã€‚
{traceback.format_exc()}
""")