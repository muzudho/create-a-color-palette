import math
import openpyxl as xl
import os
import random
import subprocess
import time
import traceback

from openpyxl.styles import PatternFill
from pathlib import Path
from tomlkit import parse as toml_parse, dumps as toml_dumps

from src.create_color_pallete import Color, ToneSystem


PATH_TO_CONFIG = './config.toml'
PATH_TO_CONTENTS = './temp/gradation.xlsx'
MAX_SCALAR = 255


class Context():
    """現在の作業状態
    """


    def __init__(self):
        self._abs_path_to_config = None
        self._config_doc_rw = None

        # ここでは、None は意志未決定、 '' は Excel アプリケーションを自動的に開かないという意志決定とします。
        self._excel_application_path = None
        self._opened_excel_process = None


    @property
    def abs_path_to_config(self):
        return self._abs_path_to_config


    @abs_path_to_config.setter
    def abs_path_to_config(self, value):
        self._abs_path_to_config = value


    @property
    def config_doc_rw(self):
        return self._config_doc_rw


    @config_doc_rw.setter
    def config_doc_rw(self, value):
        self._config_doc_rw = value


    @property
    def excel_application_path(self):
        return self._excel_application_path


    @excel_application_path.setter
    def excel_application_path(self, value):
        self._excel_application_path = value


    def is_excel_process_opened(self):
        return self._opened_excel_process is not None


    def set_opened_excel_process(self, value):
        self._opened_excel_process = value


    def terminate_opened_excel_process(self):
        self._opened_excel_process.terminate()
        self._opened_excel_process = None


def main():

    # 現在の状態を保持するオブジェクト
    context_rw = Context()

    # 設定ファイル読込
    path_of_config = Path(PATH_TO_CONFIG)
    context_rw.abs_path_to_config = path_of_config.resolve()
    print(f'🔧　read 📄［ {context_rw.abs_path_to_config} ］config file...')
    with open(context_rw.abs_path_to_config, mode='r', encoding='utf-8') as f:
        config_text = f.read()
    
    context_rw.config_doc_rw = toml_parse(config_text)

    print(f"""\
{context_rw.config_doc_rw=}
{context_rw.config_doc_rw['excel']['path']=}
""")


    while True:
        subroutine(
                context_rw=context_rw)


def subroutine(context_rw):

    print() # 空行

    if not os.path.isfile(context_rw.config_doc_rw['excel']['path']):
        while True:
            message = f"""\
🙋　Tutorial
-------------
このアプリケーションでは、 Excel アプリケーションを自動的に開いたり閉じたりしたいです。

これに同意できる方は、後述の説明を参考に Excel アプリケーションへのファイルパスを入力してください。
そうでない方は、[Ctrl] + [C] キーで強制終了していただくことができます。

Excel アプリケーションへのファイルパスの調べ方を説明します...

"""
            print(message)
            time.sleep(1)

            message = f"""\
◆ Windows 11 を使っていて、Excel をすでにインストールしている方：
    タスクバーの検索ボックスに `Excel` と入力し、
    出てきた Excel のアイコンを右クリックして［ファイルの場所を開く］をクリックしてください。
    ショートカット・アイコンが出てくるのでさらに右クリックして［ファイルの場所を開く］をクリックしてください。
    📄［EXCEL.EXE］ファイルが出てくるので右クリックして［パスのコピー］をクリックしてください。
    これでクリップボードにファイルパスがコピーされました。
    これをターミナルに貼り付けてください。
    両端にダブルクォーテーションが付いているので、ダブルクォーテーションは削除してください...

"""
            print(message)
            time.sleep(1)

            message = f"""\
◆ それ以外の方
    がんばってください。


    Example of input
    ----------------
    C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE

Input
-----
"""
            temporary_excel_application_path = input(message)
            print() # 空行

            # ワークブックを新規生成
            wb = xl.Workbook()

            # ワークシート
            ws = wb['Sheet']

            cell = ws[f'A1']
            cell.value = "ありがとうございます。 Excel ファイルを開けました。"

            cell = ws[f'A2']
            cell.value = "この画面は、プログラムの方から閉じますので、このままにしておいてください。"

            cell = ws[f'A3']
            cell.value = "引き続き、プログラムの指示に従ってください。よろしくお願いします。"

            abs_file_path_to_write = Path(PATH_TO_CONTENTS).resolve()
            print(f"""\
🔧　Save 📄［ {abs_file_path_to_write} ］file...
""")
            wb.save(abs_file_path_to_write)

            print(f"""\
🔧　Open Excel...
""")
            context_rw.set_opened_excel_process(
                subprocess.Popen([temporary_excel_application_path, abs_file_path_to_write]))    # Excel が開くことを期待
            time.sleep(1)

            message = f"""\
🙋　Tutorial
-------------
Excel アプリケーションが自動的に開かれた方は `y` を、
そうでない場合は　それ以外を入力してください。

    Example of input
    ----------------
    y

Input
-----
"""
            line = input(message)
            print() # 空行

            if line == 'y':
                context_rw.config_doc_rw['excel']['path'] = temporary_excel_application_path

                print(f"""\
🔧　Save 📄［ {context_rw.abs_path_to_config} ］config file...
""")
                with open(context_rw.abs_path_to_config, mode='w', encoding='utf-8') as f:
                    f.write(toml_dumps(context_rw.config_doc_rw))

                print(f"""\
🔧　Close Excel...
""")
                context_rw.terminate_opened_excel_process()
                time.sleep(1)
                break
                
            else:
                message = f"""\
🙋　Tutorial
-------------
もう一度、最初からやり直してください...

"""
                print(message)
                time.sleep(1)


    message = """\
Message
-------
作りたい色の数を 1 以上、常識的な数以下の整数で入力してください。

    Guide
    -----
    *   `3` - ３色
    * `100` - １００色

    Example of input
    ----------------
    7

Input
-----
"""
    line = input(message)
    number_of_color_samples = int(line)
    print() # 空行


    message = f"""\
Message
-------
彩度を 0 以上 {MAX_SCALAR} 以下の整数で入力してください。

    Guide
    -----
    *   `0` -   0 に近いほどグレー
    * `{MAX_SCALAR:3}` - {MAX_SCALAR:3} に近いほどビビッド

    Example of input
    ----------------
    {MAX_SCALAR*2//3:3}

Input
-----
"""
    line = input(message)
    saturation = int(line)
    print() # 空行


    high_brightness = MAX_SCALAR
    low_brightness = saturation
    mid_brightness = (high_brightness + low_brightness) // 2

    message = f"""\
Message
-------
明度を {low_brightness} 以上 {high_brightness} 以下の整数で入力してください。

    Guide
    -----
    *   `0` - Black
    * `100` - Dark
    * `220` - Bright
    * `{MAX_SCALAR:3}` - White

    Example of input
    ----------------
    {mid_brightness}

Input
-----
"""
    line = input(message)
    brightness = int(line)
    print() # 空行


    # ワークブックを新規生成
    wb = xl.Workbook()

    # ワークシート
    ws = wb['Sheet']

    low, high = create_tone(
            saturation=saturation,
            brightness=brightness)
    
    # 色相 [0.0, 1.0]
    cur_hue = random.uniform(0, 1)
    step_hue = 1 / number_of_color_samples
#     print(f"""\
# {step_hue=}""")


    cell = ws[f'A1']
    cell.value = "No"

    cell = ws[f'B1']
    cell.value = "色"

    cell = ws[f'C1']
    cell.value = "ウェブ・セーフ・カラー"

    # デバッグ用情報
    # cell = ws[f'A1']
    # cell.value = "色相"

    # cell = ws[f'B1']
    # cell.value = "色相種類"

    # cell = ws[f'C1']
    # cell.value = "色相内段階"


    for index, row_th in enumerate(range(2, 2 + number_of_color_samples)):

        tone_system = ToneSystem(
                low=low,
                high=high,
                hue=cur_hue)
#         print(f"""\
# {low=}
# {high=}
# {cur_hue=}""")

        color_obj = Color(tone_system.get_red(), tone_system.get_green(), tone_system.get_blue())
#         print(f"""\
# {color_obj.to_web_safe_color()=}""")
    
        web_safe_color = color_obj.to_web_safe_color()
        xl_color = web_safe_color[1:]
        try:
            pattern_fill = PatternFill(
                    patternType='solid',
                    fgColor=xl_color)
        except:
            print(f'{xl_color=}')
            raise

        # 連番
        cell = ws[f'A{row_th}']
        cell.value = index

        # 色
        cell = ws[f'B{row_th}']
        cell.fill = pattern_fill

        # ウェブ・セーフ・カラー
        cell = ws[f'C{row_th}']
        cell.value = web_safe_color

        # デバッグ情報
        # cell = ws[f'A{row_th}']
        # cell.value = cur_hue

        # cell = ws[f'B{row_th}']
        # cell.value = tone_system.get_phase_name()

        # cell = ws[f'C{row_th}']
        # cell.value = tone_system.get_value_of_hue_in_phase()

        cur_hue += step_hue
        if 1 < cur_hue:
            cur_hue -= 1


    rel_file_path_to_write = PATH_TO_CONTENTS
    path = Path(rel_file_path_to_write)
    abs_file_path_to_write = path.resolve()
    wb.save(abs_file_path_to_write)
    print(f"""\
Save 📄［ {abs_file_path_to_write} ］ file.
""")

    if context_rw.excel_application_path is None:
        message = f"""\
Message
-------
作成した結果を Excel アプリケーションで開きたいです。
できれば Excel アプリケーションのファイルパスを入力してください。
そうでなければ、そのまま Enter キーを押下してください。

    Example of input
    ----------------
    C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE

Input
-----
"""
        excel_path = input(message)
        print() # 空行
    
    else:
        excel_path = context_rw.excel_application_path


    is_successful = False
    if excel_path != '':
        print(f"""\
Attempt to start Excel.""")
        context_rw.set_opened_excel_process(
            subprocess.Popen([excel_path, abs_file_path_to_write]))    # Excel が開くことを期待


    if context_rw.is_excel_process_opened():
        print(f"""\
Please open 📄［ {abs_file_path_to_write} ］ file.
""")


    if context_rw.is_excel_process_opened():
        message = f"""\
Message
-------
自動的に開いた Excel アプリケーションを閉じたい場合は y を、
そうでない場合は　それ以外を入力してください。

    Example of input
    ----------------
    y

Input
-----
"""
        line = input(message)
        print() # 空行

        if line == 'y':
            context_rw.terminate_opened_excel_process()


def create_tone(saturation, brightness):
    """色調を１つに決めます。

    Parameters
    ----------
    saturation : int
        彩度。[0, 255] の整数
        NOTE モノクロに近づくと、標本数が多くなると、色の違いを出しにくいです。
    brightness : int
        明度
    """

    # NOTE ウェブ・セーフ・カラーは、暗い色の幅が多めに取られています。 0～255 のうち、 180 ぐらいまで暗い色です。
    # NOTE 色の標本数が多くなると、 low, high は極端にできません。変化の幅が狭まってしまいます。

    # 上限
    high = brightness
    # 下限
    low = saturation

    if 255 < high:
        raise ValueError(f'{high=} Others: {brightness=} {saturation=}')

    if low < 0:
        raise ValueError(f'{low=} Others: {brightness=} {saturation=}')


    return low, high


##########################
# MARK: コマンドから実行時
##########################

if __name__ == '__main__':
    try:
        main()

    except Exception as err:
        print(f"""\
おお、残念！　例外が投げられてしまった！
{type(err)=}  {err=}

以下はスタックトレース表示じゃ。
{traceback.format_exc()}
""")